

from decimal import Decimal, InvalidOperation
import os
import django
from django.shortcuts import get_object_or_404
import openpyxl

# from customer_credit import safe_decimal

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
django.setup()

# from decimal import Decimal
from van_management.models import Van_Routes
from datetime import datetime
from django.db import transaction
from django.utils.timezone import make_aware

from accounts.models import Customers
from client_management.models import (
    CustomerCredit,
    CustomerOutstanding,
    OutstandingAmount,
    CustomerOutstandingReport
)
from invoice_management.models import Invoice
from van_management.models import Van_Routes

from decimal import Decimal, InvalidOperation

def safe_decimal(value):
    if value is None:
        return None

    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))

    value = str(value).strip()

    if value in ("", "-", "—"):
        return None

    # remove thousand separators
    value = value.replace(",", "")

    try:
        return Decimal(value)
    except InvalidOperation:
        return None



EXCEL_FILE = "S-45 OUTSTANDING 01-05-2026 UPLOAD.xlsx"

OUT_DATE = make_aware(datetime(2026, 5, 1, 0, 0, 0))
DO_UPDATE = True        # 🔴 Set False for dry-run
# ==========================================



wb = openpyxl.load_workbook(EXCEL_FILE)
sheet = wb.active

print("=" * 90)
print("📦 BULK CUSTOMER OUTSTANDING IMPORT (OPENPYXL)")
print("📅 Outstanding Date :", OUT_DATE.date())
print("🧪 Dry Run          :", not DO_UPDATE)
print("=" * 90)

for row in sheet.iter_rows(min_row=3, values_only=True):
    customer_code = row[3]   # Column D
    name = row[4]            # Column E
    # price = row[4]           # Column F   # Column G
    excel_amount = row[5]   # Column H

    if customer_code is None or excel_amount is None:
        continue

    customer_code = str(customer_code).zfill(4)
    excel_amount = safe_decimal(excel_amount)

    try:
        customer = Customers.objects.get(custom_id=customer_code)

        adjustment = excel_amount

        print(
            f"Customer {customer_code} | "
            f"Name: {name} | "
            f"Excel: {excel_amount} | "
            f"Adjustment: {adjustment}"
        )

        if adjustment == 0:
            print("  ➖ No change needed")
            continue

        if not DO_UPDATE:
            continue

        with transaction.atomic():
            route = customer.routes
            van_route = (
                Van_Routes.objects
                .filter(routes=route)
                .select_related("van")
                .first()
            )
            salesman = van_route.van.salesman if van_route else None

            outstanding = CustomerOutstanding.objects.create(
                customer=customer,
                product_type="amount",
                created_date=OUT_DATE,
                outstanding_date=OUT_DATE,
                created_by=str(salesman.id) if salesman else ""
            )

            OutstandingAmount.objects.create(
                customer_outstanding=outstanding,
                amount=adjustment
            )

            if adjustment > 0:

                invoice = Invoice.objects.create(
                    created_date=OUT_DATE,
                    invoice_date=OUT_DATE,
                    net_taxable=adjustment,
                    vat=0,
                    discount=0,
                    amout_total=adjustment,
                    amout_recieved=0,
                    customer=customer,
                    reference_no="excel upload",
                    salesman=salesman,
                    invoice_type="credit_invoice",
                    invoice_status="non_paid"
                )

                outstanding.invoice_no = invoice.invoice_no
                outstanding.save(update_fields=["invoice_no"])
            elif adjustment < 0 :
                credit_amount = abs(adjustment)

                CustomerCredit.objects.create(
                    customer=customer,
                    amount=credit_amount,
                    source="excel_import",
                    remark="Imported customer credit from Excel",
                    salesman=customer.sales_staff,
                )

                print(f"✔ Credit added | Customer {customer_code} | Amount {credit_amount}")
            else :
               
                print(f"Zero balance | Skipped {customer_code}")

    except Customers.DoesNotExist:
        print(f"❌ Customer not found: {customer_code}")

    except Exception as e:
        print(f"❌ Error for customer {customer_code}: {e}")

print("=" * 90)
print("✅ BULK IMPORT COMPLETED")
print("=" * 90)


# EXCEL_FILE = "S-45 CUSTOMER NAME CHANGE.xlsx"

# wb = openpyxl.load_workbook(EXCEL_FILE)
# sheet = wb.active

# updated_count = 0
# not_found = []

# print("Starting update process...")

# # Wrap in a transaction for speed and data integrity
# try:
#     with transaction.atomic():
#         # Start from row 3 as per your logic
#         for row in sheet.iter_rows(min_row=3, values_only=True):
#             customer_code = row[3]      # Column H (Index 7)
#             new_name = row[6]           # Column I (Index 8)
#             new_building = row[4]      # Column K (Index 10)
#             house_no = row[5]

#             # Skip if the customer_code is empty
#             if not customer_code:
#                 continue

#             # Assuming 'custom_id' is the field that matches your Excel Customer Code
#             # Change 'custom_id' to 'customer_id' if that is your unique code field
#             try:
#                 customer = Customers.objects.get(custom_id=customer_code)
#                 customer.customer_name = new_name
#                 customer.building_name = new_building
#                 customer.door_house_no = house_no
#                 customer.save()
#                 updated_count += 1
#                 print("updated count: ",updated_count)
#             except Customers.DoesNotExist:
#                 not_found.append(customer_code)
#             except Exception as e:
#                 print(f"Error updating {customer_code}: {e}")

#     print("--- Update Summary ---")
#     print(f"Successfully updated: {updated_count} customers.")
#     if not_found:
#         print(f"Total not found in DB: {len(not_found)}")
#         print(f"Sample of missing codes: {not_found[:5]}")

# except Exception as e:
#     print(f"Critical Transaction Error: {e}")


import os
import django
from datetime import date, datetime

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
django.setup()

from django.db import transaction

from invoice_management.models import Invoice
from client_management.models import CustomerCredit, CustomerOutstanding, OutstandingAmount


# # --------------------------------------------------
# # CONFIG
# # --------------------------------------------------
# REFERENCE = "excel upload"
# ROUTE_NAME = "S-03"


# # --------------------------------------------------
# # MAIN FUNCTION
# # --------------------------------------------------
# def delete_excel_upload_invoices():
#     OUT_DATE = make_aware(datetime(2026, 3, 17, 0, 0, 0))

#     summary = {
#         "invoices": 0,
#         "outstanding": 0,
#         "outstanding_amount": 0,
#     }

#     with transaction.atomic():

#         print(f"\n🔍 Deleting Excel Upload Data")
#         print(f"Reference: {REFERENCE}")
#         print(f"Date: {OUT_DATE}")

#         # -------------------------------
#         # Get Invoices
#         # -------------------------------
#         invoices = Invoice.objects.filter(
#             reference_no=REFERENCE,   # change if field is reference_no
#             invoice_date__date=OUT_DATE,
#             customer__routes__route_name="S-03",
#             is_deleted=False
#         )

#         if not invoices.exists():
#             print("❌ No invoices found")
#             return summary

#         # -------------------------------
#         # Loop invoices
#         # -------------------------------
#         for inv in invoices:
#             invoice_no = inv.invoice_no
#             customer = inv.customer

#             print(f"\n🧾 Invoice: {invoice_no}")

#             # -------------------------------
#             # CustomerOutstanding
#             # -------------------------------
#             outstanding_qs = CustomerOutstanding.objects.filter(
#                 customer=customer,
#                 invoice_no=invoice_no
#             )

#             for oh in outstanding_qs:
#                 oa_qs = OutstandingAmount.objects.filter(
#                     customer_outstanding=oh
#                 )

#                 count = oa_qs.count()
#                 oa_qs.delete()
#                 summary["outstanding_amount"] += count

#             count = outstanding_qs.count()
#             outstanding_qs.delete()
#             summary["outstanding"] += count

#             print(f"  ✔ OutstandingAmount deleted: {summary['outstanding_amount']}")
#             print(f"  ✔ CustomerOutstanding deleted: {count}")

#             # -------------------------------
#             # Invoice
#             # -------------------------------
#             inv.delete()
#             summary["invoices"] += 1

#             print("  ✔ Invoice deleted")

#     print("\n🎉 DONE")
#     for k, v in summary.items():
#         print(f"{k}: {v}")

#     return summary


# # --------------------------------------------------
# # RUN
# # --------------------------------------------------
# if __name__ == "__main__":
#     delete_excel_upload_invoices()